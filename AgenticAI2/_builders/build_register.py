#!/usr/bin/env python3
"""
_builders/build_register.py
===========================
Generates 00_Program/Assumptions_and_Decisions_Register.xlsx

This workbook is a governance artifact, not a data dump. It is the document you
hand a client sponsor when they ask "what did you assume, what is unresolved, and
what might break?" — the three questions that decide whether a training
programme's outputs can be trusted downstream.

Six sheets:
    Cover              orientation and how to read it
    Assumptions        every number we chose rather than measured
    Open Questions     specification gaps + curriculum gaps, with severity
    Version Risks      technical surfaces that can go stale, with confidence
    Evidence Register  every factual claim, classified by how we know it
    Lab Inventory      what is built, what runs, what is outstanding
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "00_Program" / "Assumptions_and_Decisions_Register.xlsx"

NAVY, DEEP, TEAL, MINT, GOLD = "21295C", "065A82", "1C7293", "16A0A0", "E0A800"
WASH, LINE = "F2F6F9", "D7E1EA"

HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10.5, color="3C4257")
MONO_FONT = Font(name="Consolas", size=10, color="3C4257")
TITLE_FONT = Font(name="Cambria", size=20, bold=True, color=NAVY)
SUB_FONT = Font(name="Calibri", size=11, italic=True, color="6B7280")

SEV_FILL = {
    "HIGH": PatternFill("solid", fgColor="F8D7DA"),
    "MEDIUM": PatternFill("solid", fgColor="FDF6E3"),
    "LOW": PatternFill("solid", fgColor="E8F5F3"),
    "CLOSED": PatternFill("solid", fgColor="E3F1E8"),
    "OPEN": PatternFill("solid", fgColor="FDF6E3"),
    "BLOCKED": PatternFill("solid", fgColor="F8D7DA"),
    "BUILT": PatternFill("solid", fgColor="E3F1E8"),
    "NOT BUILT": PatternFill("solid", fgColor="F8D7DA"),
    "DESIGN ONLY": PatternFill("solid", fgColor="FDF6E3"),
}
THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sheet(wb: Workbook, name: str, title: str, subtitle: str,
          headers: list[str], widths: list[int], rows: list[list],
          sev_col: int | None = None, freeze: str = "A5") -> None:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUB_FONT
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    for i, (header, width) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
        c = ws.cell(row=4, column=i, value=header)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    for r, row in enumerate(rows, start=5):
        for i, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=value)
            c.font = MONO_FONT if i == 1 else BODY_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if sev_col and i == sev_col:
                fill = SEV_FILL.get(str(value).upper())
                if fill:
                    c.fill = fill
                    c.font = Font(name="Calibri", size=10.5, bold=True, color="3C4257")
        ws.row_dimensions[r].height = None

    ref = f"A4:{get_column_letter(len(headers))}{4 + len(rows)}"
    table = Table(displayName=f"tbl{name.replace(' ', '')}", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = freeze


wb = Workbook()

# ------------------------------------------------------------------ Cover --
cover = wb.active
cover.title = "Cover"
cover.sheet_view.showGridLines = False
cover.column_dimensions["A"].width = 3
cover.column_dimensions["B"].width = 108

lines = [
    ("Accenture Batch 1 — Agentic AI Foundation", TITLE_FONT, 32),
    ("Assumptions & Decisions Register", Font(name="Cambria", size=15, color=DEEP), 24),
    ("", BODY_FONT, 10),
    ("WHAT THIS IS", Font(name="Calibri", size=11, bold=True, color=NAVY), 18),
    ("The governance record for the courseware. It answers the three questions a client sponsor asks "
     "before trusting a programme's outputs: what did you assume, what is unresolved, and what might break?",
     BODY_FONT, 30),
    ("", BODY_FONT, 8),
    ("HOW TO READ IT", Font(name="Calibri", size=11, bold=True, color=NAVY), 18),
    ("Assumptions — every number chosen rather than measured. None of these is a finding. Each names the "
     "experiment that would replace it.", BODY_FONT, 26),
    ("Open Questions — specification gaps (S) belong to the client; curriculum gaps (G) belong to us. "
     "Ranked by remediation priority.", BODY_FONT, 26),
    ("Version Risks — technical surfaces that can go stale, with a confidence rating and a re-check action.",
     BODY_FONT, 20),
    ("Evidence Register — every factual claim in the courseware, classified as measured, client-specified, "
     "assumption, or courseware invention.", BODY_FONT, 26),
    ("Lab Inventory — what is built, what has been executed, and what is outstanding.", BODY_FONT, 20),
    ("", BODY_FONT, 8),
    ("THE MEASUREMENT POSITION", Font(name="Calibri", size=11, bold=True, color=NAVY), 18),
    ("This courseware makes NO productivity claim and asserts NO automation target. Both are treated as "
     "measurement problems requiring a baseline frozen on the client's own payment file before training "
     "begins. Match rate and straight-through rate are different numbers; conflating them is the most "
     "common way an O2C automation business case is overstated.", BODY_FONT, 46),
    ("", BODY_FONT, 8),
    ("SOURCE", Font(name="Calibri", size=11, bold=True, color=NAVY), 18),
    ("Reviewed against Accenture_Batch_1_Detailed.pdf. Figures reproduce from a real run of the packaged "
     "labs; see the Evidence Register for the backend that produced each one.", BODY_FONT, 26),
]
row = 2
for text, font, height in lines:
    cover.cell(row=row, column=2, value=text).font = font
    cover.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    cover.row_dimensions[row].height = height
    row += 1

# ------------------------------------------------------------ Assumptions --
sheet(
    wb, "Assumptions",
    "Assumptions", "Numbers we CHOSE rather than measured. Each names the experiment that would replace it.",
    ["ID", "Assumption", "Value", "Where it is used", "Why this value", "Experiment that would replace it", "Owner"],
    [8, 34, 14, 26, 34, 44, 16],
    [
        ["A-01", "Auto write-off tolerance", "$10.00", "Day 1 Lab 4/5, Capstone",
         "Client-specified in the source document (Example C). Carried as given.",
         "None needed — client-specified. But test tiered/percentage tolerance if invoice values span orders of magnitude.",
         "Client (O2C owner)"],
        ["A-02", "Retrieval relevance threshold", "0.92 cosine distance", "Day 2 Lab 3, Capstone",
         "Chosen against this 10-document corpus with the offline lexical embedder. Not validated.",
         "Label a set of query/document pairs; measure precision and recall at several thresholds; pick where cost of a miss equals cost of a false block.",
         "Delivery team"],
        ["A-03", "Confidence floor for auto-dispute", "0.60", "Day 2 Lab 5, Capstone",
         "Asserted for courseware. The model's confidence is not a calibrated probability.",
         "Label real deductions; measure error rate by confidence band; set floor where cost of a wrong code equals cost of an unnecessary review.",
         "Delivery team"],
        ["A-04", "Chunk minimum size", "120 characters", "Day 2 Lab 1",
         "Prevents a two-line heading becoming its own record in these documents.",
         "Re-test on real remittance PDFs, which are longer and more tabular than the seed set.",
         "Delivery team"],
        ["A-05", "Chunk overlap", "1 line", "Day 2 Lab 1",
         "Carries a sentence spanning a boundary into both chunks.",
         "Measure retrieval recall for evidence that straddles a chunk edge.",
         "Delivery team"],
        ["A-06", "Retry limit on model extraction", "2 attempts", "Day 2 Lab 4",
         "Retrying forever on a persistently malformed response burns budget and hides the fault.",
         "Measure how often attempt 2 succeeds where attempt 1 failed. If near zero, drop to 1.",
         "Delivery team"],
        ["A-07", "Deduction owning team and SLA", "Quality 10d, Pricing 5d, etc.", "Day 2 Labs 2/5, webapps",
         "COURSEWARE INVENTION. Not in the source document. Added so learners see that a code alone is not actionable.",
         "Replace wholesale with the client's real ownership matrix before any client-facing use.",
         "Client (O2C owner)"],
        ["A-08", "UAC vs UIC test", "Payer identifiability", "Day 1 Lab 4, Capstone",
         "Source document defines UAC two ways. We implement the Example D reading.",
         "Client decision — see Open Question S2.",
         "Client (O2C owner)"],
        ["A-09", "Money represented as float", "Python float", "All labs",
         "Readability on a projector. Explicitly flagged in Day 1 Lab 3 and slide 11.",
         "None — production must use decimal.Decimal. This is a courseware simplification, not a design position.",
         "Delivery team"],
        ["A-10", "Offline embedder dimension", "256", "shared/foundry_client.py",
         "Arbitrary; large enough to avoid hash collisions on this corpus.",
         "Irrelevant in production — Azure embeddings replace it entirely.",
         "Delivery team"],
    ],
    freeze="A5",
)

# --------------------------------------------------------- Open Questions --
sheet(
    wb, "Open Questions",
    "Open Questions", "S = client's specification to decide. G = our curriculum to fix. Ranked by remediation priority.",
    ["ID", "Title", "Type", "Severity", "Status", "Evidence in the build", "Recommendation", "Priority"],
    [8, 28, 12, 12, 12, 40, 42, 9],
    [
        ["G3", "Human-in-the-loop not taught", "Curriculum", "HIGH", "CLOSED",
         "CLOSED by Day 3 Lab 6. interrupt() plus a durable SqliteSaver checkpointer; QUERY suspends, persists and resumes from a freshly built graph object with analyst attribution intact. All three decision paths verified.",
         "No further action. Confirm langgraph-checkpoint-sqlite is installed - verify_environment.py checks it. Capstone acceptance criteria 5 and 6 are now reachable.",
         9],
        ["S3", "One-to-many cash application undefined", "Specification", "HIGH", "OPEN",
         "BNK-1008 remits $15,000 covering INV-1102 ($9,000) and INV-1103 ($6,000). Priority 5 fires, matches the first invoice, reports a false $6,000 overpayment.",
         "Client decision. Needs an applications list in state plus a matching end state, or an explicit scope exclusion.",
         2],
        ["S1", "No end state for an overpayment", "Specification", "HIGH", "OPEN",
         "BNK-1010 pays $12,000 against $11,000. Fits none of the six states. Routed to QUERY with 'SPECIFICATION GAP' in the trace.",
         "Client decision: add a seventh state, or a documented rule folding it into UAC. Do not invent one silently.",
         3],
        ["S5", "No idempotency or replay position", "Specification", "HIGH", "MITIGATED",
         "Nothing in the spec covers batch re-runs or crash-after-post. Build adds content-derived idempotency keys on create_dispute and on ingestion.",
         "Confirm the approach with the client's ERP team; idempotency semantics must match the real posting API.",
         4],
        ["G4", "Scaffolding jump to the capstone", "Curriculum", "MEDIUM", "OPEN",
         "Day 2 Lab 5 ends with a 13-node in-memory graph over 10 transactions. Capstone expects batch handling, persistence and a review UI.",
         "Short bridging lab on batch orchestration and durable state, or an explicit statement that the capstone is a guided build.",
         5],
        ["G2", "MCP transport not taught", "Curriculum", "MEDIUM", "PARTIAL",
         "Day 2 Lab 2 builds the tool boundary fully — typed schemas, read/write permissions, refusals, audit. No actual MCP client/server over a transport.",
         "Add a ~45-minute MCP lab to Day 3, or state to the client that MCP is architectural context rather than a hands-on outcome.",
         6],
        ["S2", "UAC defined two ways", "Specification", "MEDIUM", "MITIGATED",
         "End-state table says 'no invoice AND no remittance'; Example D says 'customer identified, invoice unknown'. Build implements Example D.",
         "Confirm with the client which test governs. It decides whether Cash Application or Treasury owns the item.",
         7],
        ["S4", "No currency or FX handling", "Specification", "MEDIUM", "OPEN",
         "Every amount in the spec is USD. Not implemented; raised as a Day 1 Lab 5 stretch goal.",
         "Scoping decision. Blocks multi-region rollout only. Needs a rate source, rate-date convention, and FX variance treated separately from a deduction.",
         8],
        ["G1", "Structured-output validation absent", "Curriculum", "HIGH", "CLOSED",
         "Days 1-3 as scoped parsed JSON defensively. Day 2 Lab 4 now enforces a Pydantic contract plus a verbatim-citation grounding check; 5 bad payloads rejected in class.",
         "No further action. Verify the lab runs on the delivery machine.",
         9],
    ],
    freeze="A5",
)

# ----------------------------------------------------------- Version Risks --
sheet(
    wb, "Version Risks",
    "Version Risks", "Technical surfaces that can go stale. Re-run verify_environment.py on the delivery machine.",
    ["ID", "Surface", "Risk", "Confidence", "Mitigation in the build", "Action before delivery"],
    [8, 30, 40, 12, 42, 38],
    [
        ["A1", "azure-ai-projects client accessor",
         "The method to obtain a chat client from AIProjectClient has appeared in more than one shape across the preview line and GA.",
         "LOW",
         "foundry_client.py PROBES known accessor shapes and raises with the installed version if none work. Path A is the teaching path.",
         "Verify against current Microsoft Learn. Do NOT present a single Path B signature as settled."],
        ["A2", "openai AzureOpenAI client", "Stable surface. api_version string still pins server behaviour.", "HIGH",
         "Pinned openai>=1.55,<2. api_version set explicitly in .env.",
         "Confirm the api_version is one your resource accepts."],
        ["A3", "model= takes the DEPLOYMENT name", "Passing the model family name instead is the top day-one support ticket.", "HIGH",
         "Called out in Day 1 Lab 2, deck slide 14 and the speaker notes.",
         "Say it out loud in Lab 2."],
        ["A4", "LangGraph StateGraph / conditional edges", "Core API, stable across the versions in use.", "HIGH",
         "Built and executed against 0.2.60+ and 1.2.x.", "None."],
        ["A5", "LangGraph interrupt / Command (HITL)",
         "Post-0.2 API. Now exercised directly by Day 3 Lab 6.", "HIGH",
         "Built and executed against the installed version. Lab 6 verifies the import at start-up and falls back to MemorySaver with an explicit warning if SqliteSaver is absent.",
         "Confirm langgraph-checkpoint-sqlite is installed, or Lab 6 Step 6 cannot demonstrate durability."],
        ["A6", "chromadb client and embedding-function API", "Changed shape at the 0.5 to 1.0 boundary.", "MEDIUM",
         "Labs pass embeddings EXPLICITLY and never use Chroma's default embedding function. Built and tested against 1.5.9.",
         "Confirm the installed version imports and PersistentClient constructs."],
        ["A7", "Chroma default embedding function",
         "Downloads a local ONNX model on first use: surprise network call, unpinned model, different vector space from production.",
         "N/A", "Avoided by design. Explained in Day 2 Lab 1 Step 5.",
         "Do not reintroduce it."],
        ["A8", "Pydantic v2 field_validator", "v1 @validator syntax is incompatible.", "HIGH",
         "Pinned pydantic>=2.9,<3.", "None."],
        ["A9", "draw_mermaid() on compiled graphs", "Availability varies by LangGraph version.", "MEDIUM",
         "Every call wrapped with an ASCII fallback. No lab depends on it.", "None."],
        ["A10", "Relative CHROMA_PERSIST_DIR",
         "A relative path resolved against the working directory put the collection where the web app could not find it — empty collection, no error.",
         "HIGH", "config.py now anchors relative paths to the repository root.",
         "None. Fixed and verified."],
    ],
    freeze="A5",
)

# ------------------------------------------------------- Evidence Register --
sheet(
    wb, "Evidence Register",
    "Evidence Register", "Every factual claim, classified by how we know it. Any figure quoted externally must carry its backend.",
    ["ID", "Claim", "Classification", "Basis", "Caveat when quoting externally"],
    [8, 40, 22, 42, 42],
    [
        ["E-01", "Six end states, six priority rules, codes D01-D05", "Client-specified",
         "Taken directly from Accenture_Batch_1_Detailed.pdf.",
         "Not independently verified against a live Accenture process."],
        ["E-02", "$10 auto write-off tolerance", "Client-specified", "Stated in the source document, Example C.", "None."],
        ["E-03", "Owning team and SLA per deduction code", "COURSEWARE INVENTION",
         "Not in the source document. Added so learners see routing, not just classification.",
         "MUST be replaced with the client's real ownership matrix before client-facing use."],
        ["E-04", "Day 1 baseline: 70% match rate, 40% straight-through", "Measured, reproducible",
         "Computed from the 10-row seed file. Enforced as a known-answer test in verify_environment.py.",
         "Seed data, not client data. Re-measure on the client's payment file."],
        ["E-05", "Day 2 delta: straight-through unchanged, 1/10 coded", "Measured on the offline stub",
         "Reproducible offline. Day 2 Lab 5 prints it.",
         "Offline stub with a LEXICAL embedder. Figures on a live model will differ."],
        ["E-06", "Relevance threshold 0.92", "Assumption", "Chosen against this corpus and embedder.", "See Assumption A-02."],
        ["E-07", "Confidence floor 0.60", "Assumption", "Asserted for courseware.", "See Assumption A-03."],
        ["E-08", "Customer names (Acme, Globex, Umbrella Health...)", "Synthetic",
         "Deliberately fictional. No real customer data anywhere in the package.", "None."],
        ["E-09", "Model endpoint latency", "Learner-measured", "Day 2 Lab 2 has learners measure their own endpoint.",
         "No latency figure is claimed by the courseware."],
        ["E-10", "Any productivity or automation-uplift percentage", "NOT CLAIMED",
         "Deliberately absent. Treated as a measurement problem requiring a frozen baseline.",
         "If asked for a number, the correct answer is 'give me a week with your payment file'."],
        ["E-11", "Offline embedder is lexical, not semantic", "Measured limitation",
         "Hashed bag-of-words. 'Short paid' and 'underpaid' share no tokens and do not retrieve each other.",
         "Never quote retrieval QUALITY from this backend."],
        ["E-12", "BNK-1009 correctly returns UNKNOWN", "Measured, reproducible",
         "Remittance states no reason. Grounding + closed code set produce UNKNOWN, routed to QUERY.",
         "Demonstrates a LOWER automation rate and a HIGHER correctness rate. Report both."],
    ],
    freeze="A5",
)

# --------------------------------------------------------- Lab Inventory ---
sheet(
    wb, "Lab Inventory",
    "Lab Inventory", "Built = artifact exists. Executed = run to completion during the build.",
    ["Day", "Artifact", "Status", "Executed", "Notes"],
    [16, 40, 14, 12, 48],
    [
        ["Shared", "config / telemetry / foundry_client / vectorstore", "BUILT", "yes",
         "Four seams. An SDK breaking change costs one file, not fifteen labs."],
        ["Shared", "Seed data: 10 bank rows, 9 AR items, 5 remittances", "BUILT", "yes",
         "Engineered so every priority rule and every end state fires, including the undefined ones."],
        ["00_Program", "requirements / .env.example / verify_environment.py", "BUILT", "yes",
         "Verifier includes a known-answer test on the Day 1 baseline."],
        ["00_Program", "Version risk + gap registers, this workbook", "BUILT", "n/a", ""],
        ["00_Program", "LAB_ENVIRONMENT_GUIDE.md + Word version", "BUILT", "n/a",
         "Four run options: local VS Code, local Jupyter, GitHub Codespaces, Azure ML compute instance. Decision matrix, troubleshooting, trainer checklist."],
        ["root", ".devcontainer/ + .vscode/ + setup.sh + setup.ps1 + environment.yml", "BUILT", "partial",
         "devcontainer.json and all .vscode JSON validated; setup.sh logic exercised. Codespaces and Azure ML flows NOT executed end to end - see the version notice in guide section 6."],
        ["00_Program", "Day 0 kickoff deck (12 slides) + facilitation guide", "BUILT", "n/a",
         "Orientation, environment setup, how the lab files work, honest build status. 57 min. Deliver the afternoon BEFORE Day 1."],
        ["_builders", "validate_labs.py — lab validation harness", "BUILT", "yes",
         "1,538 checks across 10 solutions, 10 starters, 20 notebooks: structure, derivation, notebook schema, execution. Wired into build_all.sh."],
        ["Day 1", "Deck — 18 slides, 21,203 note chars", "BUILT", "n/a", "Schema-validated. Visual QA incomplete — scroll through before delivery."],
        ["Day 1", "Facilitation guide (Word)", "BUILT", "n/a", "Generated from content_day1.js. Cannot drift from the deck."],
        ["Day 1", "Labs 1-5 (solutions + starters + notebooks)", "BUILT", "yes", "5/5 solutions pass. Starters halt at blank 1 as designed."],
        ["Day 1", "Webapp — reconciliation console", "BUILT", "partial", "Syntax and graph import verified; not exercised in a browser."],
        ["Day 2", "Deck — 13 slides, 14,724 note chars", "BUILT", "n/a", "Schema-validated."],
        ["Day 2", "Facilitation guide (Word)", "BUILT", "n/a", "Generated from content_day2.js."],
        ["Day 2", "Labs 1-5 (solutions + starters + notebooks)", "BUILT", "yes", "5/5 pass against real ChromaDB 1.5.9. Two blanks each after a validation pass flagged single-blank labs as thin."],
        ["Day 2", "Webapp — remittance explorer", "BUILT", "partial", "Syntax and collection access verified; not exercised in a browser."],
        ["Day 3", "Deck — 13 slides, 80 min", "BUILT", "n/a",
         "Updated after the labs landed: lab sequence now lists six labs and gap G3 is marked closed."],
        ["Day 3", "Facilitation guide (Word)", "BUILT", "n/a", "Generated from content_day3.js."],
        ["Day 3", "Labs 1-6 (solutions + starters + notebooks)", "BUILT", "yes",
         "6/6 pass. Input gates (catch 83%, FP 25%), output redaction, transition auditing, secured 16-node pipeline, 6/6 scenario matrix, and HITL."],
        ["Day 3", "Webapp — guardrail sandbox", "NOT BUILT", "no", "Optional. The scenario matrix in Lab 5 covers the teaching need."],
        ["Capstone", "Deck — 12 slides", "DESIGN ONLY", "n/a", "Design-locked ahead of the build."],
        ["Capstone", "Facilitation guide (Word)", "DESIGN ONLY", "n/a", "Same caveat as the deck."],
        ["Capstone", "src/ package (domain, security, pipeline, batch, cli)", "BUILT", "yes",
         "19-node graph with checkpointing and HITL. Reproduces the frozen Day 1-3 baseline: 40% straight-through, 0 failures."],
        ["Capstone", "tests/test_acceptance.py — the 7 criteria", "BUILT", "yes",
         "7/7 PASS. Criterion 2 doubles as the anti-drift check between the lab code and the package."],
        ["Capstone", "webapp — analyst review console", "BUILT", "partial",
         "Syntax and imports verified; not exercised in a browser. Four exception queues plus a Decide tab that resumes the graph."],
        ["Capstone", "docs/BUILD_GUIDE.md", "BUILT", "n/a",
         "Package layout, run commands, acceptance criteria, measured results, open items, operational realities."],
        ["Day 3", "Lab 6 — HITL (closes gap G3)", "BUILT", "yes",
         "interrupt() + SqliteSaver. Suspend, persist, resume from a fresh graph object with attribution. Was the highest-priority outstanding item."],
    ],
    sev_col=3, freeze="A5",
)

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT.relative_to(ROOT)} — {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")
